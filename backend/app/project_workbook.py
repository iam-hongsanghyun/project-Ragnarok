"""Lossless, round-trippable PyPSA *project* workbook (read + write).

This module reads and writes the **exact xlsx layout** that the Ragnarok
frontend's ``buildProjectWorkbook`` / ``parseProjectWorkbook`` use
(``frontend/.../lib/workbook/workbook.ts``). Producing that layout on the
server means:

* **Export** — a stored run's JSON bundle is rendered to an xlsx the frontend
  can re-open with full analytics, with *no information dropped* (unlike the
  previous ``OUT_*`` inspection sheets, which truncated names and could not be
  re-imported).
* **Import** — an uploaded project xlsx is parsed straight back into a run
  bundle (``{model, scenario, options, result}``) and handed to
  :func:`run_store.store_run`, so it lands in History like any solved run.

Layout (must stay in lock-step with the frontend):

* One sheet per model component (``generators``, ``buses``, …). Solved
  *static* outputs (e.g. ``p_nom_opt``) are merged in as extra columns; the
  importer splits them back out by the schema's input/output classification.
* Input time-series sheets (``generators-p_max_pu``, ``loads-p_set``, …) and
  config sheets (``RAGNAROK_Scenarios`` …) are copied verbatim.
* Output time-series sheets named ``<list>-<attr>`` (``generators-p``,
  ``storage_units-state_of_charge`` …) — never an ``OUT_`` prefix, so the names
  stay inside Excel's 31-char limit and parse cleanly.
* ``RAGNAROK_ResultMeta`` / ``RAGNAROK_Constraints`` / ``RAGNAROK_RunState`` /
  ``RAGNAROK_Settings`` / ``RAGNAROK_PluginAnalytics`` metadata sheets. Long
  JSON payloads are chunked across rows (Excel caps a cell at 32 767 chars).

The input/output column classification comes from PyPSA's own component schema
(``pypsa_schema``) — the same source the frontend schema is generated from — so
the two implementations cannot drift apart.
"""
from __future__ import annotations

import json
import logging
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..pypsa import pypsa_schema as ps

logger = logging.getLogger("pypsa_gui.project_workbook")

# ── Sheet names — MUST match the frontend constants exactly ──────────────────
RESULT_META_SHEET = "RAGNAROK_ResultMeta"
PLUGIN_ANALYTICS_SHEET = "RAGNAROK_PluginAnalytics"
PLUGIN_CONFIGS_SHEET = "RAGNAROK_PluginConfigs"
SETTINGS_SHEET = "RAGNAROK_Settings"
CONSTRAINTS_SHEET = "RAGNAROK_Constraints"
RUN_STATE_SHEET = "RAGNAROK_RunState"
RUN_HISTORY_SHEET = "RAGNAROK_RunHistory"
PROVENANCE_SHEET = "RAGNAROK_Provenance"
# Human-readable landing sheet (KPIs, constraints, run settings). Display-only —
# carries no data the importer needs, so it's in _META_SHEETS (skipped on read).
SUMMARY_SHEET = "RAGNAROK_Summary"

# The COMPLETE run bundle (model + scenario + options + the full derived result)
# is embedded here as chunked JSON. This is the canonical, lossless payload: on
# import it is read back verbatim, so an imported run is byte-for-byte the run
# that was exported — no field is reconstructed or lost. The readable component
# / series / metadata sheets are kept alongside it purely for human inspection
# (and frontend-parser compatibility).
BUNDLE_SHEET = "RAGNAROK_Bundle"

# Sheets the importer must NOT treat as model/output data.
_META_SHEETS = {
    RESULT_META_SHEET,
    PLUGIN_ANALYTICS_SHEET,
    PLUGIN_CONFIGS_SHEET,
    SETTINGS_SHEET,
    CONSTRAINTS_SHEET,
    RUN_STATE_SHEET,
    RUN_HISTORY_SHEET,
    PROVENANCE_SHEET,
    BUNDLE_SHEET,
    SUMMARY_SHEET,
}

# JSON metadata is chunked at this many chars/row (Excel cell limit is 32 767).
_MAX_CELL_CHARS = 30_000
_EXCEL_SHEET_LIMIT = 31

# result-level keys serialised into RAGNAROK_ResultMeta (chunked JSON).
_RESULT_META_KEYS = ("runMeta", "pathway", "rolling", "co2Shadow", "narrative")


# ── Schema helpers (PyPSA-schema driven; mirror the frontend) ────────────────
def _output_static_attrs(sheet: str) -> set[str]:
    """Output attributes stored as a static column (e.g. ``p_nom_opt``)."""
    comp = ps.component_schema(sheet)
    if comp is None:
        return set()
    return set(comp.get("output_attributes", [])) - set(comp.get("temporal_attributes", []))


def _output_series_attrs(sheet: str) -> set[str]:
    """Output attributes stored as a time series (e.g. ``p``, ``state_of_charge``)."""
    comp = ps.component_schema(sheet)
    if comp is None:
        return set()
    return set(comp.get("output_attributes", [])) & set(comp.get("temporal_attributes", []))


def _chunks(text: str) -> list[str]:
    if len(text) <= _MAX_CELL_CHARS:
        return [text]
    return [text[i : i + _MAX_CELL_CHARS] for i in range(0, len(text), _MAX_CELL_CHARS)]


# ─────────────────────────────────────────────────────────────────────────────
#  WRITE  —  bundle → xlsx bytes
# ─────────────────────────────────────────────────────────────────────────────
def _ordered_columns(rows: list[dict[str, Any]]) -> list[str]:
    """Union of keys across rows, in first-seen order (stable column order)."""
    cols: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                cols.append(key)
    return cols


def _write_rows(
    writer: "pd.ExcelWriter",
    name: str,
    rows: list[dict[str, Any]] | None,
    used: set[str],
) -> bool:
    """Write ``rows`` as a sheet ``name``; return True if a sheet was written.

    Skips empty input. Truncates the name to Excel's 31-char limit and dedupes
    collisions (output series names are <= 29 chars, so this is a safety net).
    """
    if not rows:
        return False
    sheet = name[:_EXCEL_SHEET_LIMIT]
    suffix = 1
    while sheet in used:
        tail = f"_{suffix}"
        sheet = name[: _EXCEL_SHEET_LIMIT - len(tail)] + tail
        suffix += 1
    if sheet != name:
        logger.warning("Sheet name %r truncated/deduped to %r", name, sheet)
    used.add(sheet)
    df = pd.DataFrame(rows, columns=_ordered_columns(rows))
    df.to_excel(writer, sheet_name=sheet, index=False)
    return True


def _merge_static_outputs(
    sheet: str,
    rows: list[dict[str, Any]],
    static_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Strip output-static columns from input rows, then merge solved values in.

    Mirrors the frontend writer: a component row carries only its input columns
    plus the *solved* static outputs (``p_nom_opt`` …) from ``outputs.static``.
    Components that exist only in the outputs (e.g. auto-added load shedding)
    are appended.
    """
    out_attrs = _output_static_attrs(sheet)
    merged: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    for row in rows:
        stripped = {k: v for k, v in row.items() if k not in out_attrs}
        name = row.get("name")
        if name is not None:
            seen_names.add(str(name))
            solved = static_map.get(str(name))
            if solved:
                stripped.update(solved)
        merged.append(stripped)
    for name, attrs in static_map.items():
        if name not in seen_names:
            merged.append({"name": name, **attrs})
    return merged


def _result_meta_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key in _RESULT_META_KEYS:
        value = result.get(key)
        if value is None or value == [] or value == {}:
            continue
        for part, chunk in enumerate(_chunks(json.dumps(value))):
            rows.append({"key": key, "part": part, "json": chunk})
    return rows


def _plugin_analytics_rows(plugin: dict[str, Any] | None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not isinstance(plugin, dict):
        return rows
    for module_id, entry in plugin.items():
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name", module_id))
        for field in ("ui", "data"):
            payload = json.dumps(entry.get(field, {}))
            for part, chunk in enumerate(_chunks(payload)):
                rows.append(
                    {"moduleId": module_id, "name": name, "field": field, "part": part, "value": chunk}
                )
    return rows


def _kv_rows(pairs: list[tuple[str, Any]]) -> list[dict[str, Any]]:
    return [{"key": k, "value": v} for k, v in pairs if v is not None]


# ── Human-readable Summary sheet ──────────────────────────────────────────────
def _summary_rows(bundle: dict[str, Any]) -> list[dict[str, Any]]:
    """A skimmable {Section, Item, Value} overview for the Summary landing sheet.

    Pulls from the already-human-formatted ``result.summary`` KPIs (values carry
    units), the run window/meta, the active constraints (their labels are plain
    language), and the headline settings — so opening the workbook lands on
    something a person can read, not the raw ``network`` sheet.
    """
    result = bundle.get("result") if isinstance(bundle.get("result"), dict) else {}
    scenario = bundle.get("scenario") if isinstance(bundle.get("scenario"), dict) else {}
    options = bundle.get("options") if isinstance(bundle.get("options"), dict) else {}
    run_meta = result.get("runMeta") if isinstance(result.get("runMeta"), dict) else {}
    rolling = run_meta.get("rolling") if isinstance(run_meta.get("rolling"), dict) else {}

    rows: list[dict[str, Any]] = []
    pending_section: str | None = None

    def section(name: str) -> None:
        nonlocal pending_section
        pending_section = name

    def add(item: str, value: Any) -> None:
        nonlocal pending_section
        if value is None or value == "":
            return
        rows.append({"Section": pending_section or "", "Item": item, "Value": value})
        pending_section = None

    label = bundle.get("label") or options.get("scenarioLabel") or options.get("filename") or "Run"
    section("Overview")
    add("Run", label)
    add("Source file", options.get("filename"))
    add("Backend", options.get("backend"))
    add("Solver", options.get("solverType"))

    section("Window")
    snaps = run_meta.get("snapshotCount")
    if snaps is None and options.get("snapshotStart") is not None and options.get("snapshotEnd") is not None:
        snaps = options.get("snapshotEnd") - options.get("snapshotStart")
    add("Snapshots", snaps)
    weight = run_meta.get("snapshotWeight") or options.get("snapshotWeight")
    add("Resolution", f"{weight}h" if weight is not None else None)
    add("Modelled hours", run_meta.get("modeledHours"))
    add("Planning mode", run_meta.get("planningMode"))
    if rolling.get("enabled"):
        add(
            "Rolling horizon",
            f"horizon {rolling.get('horizonSnapshots')} / overlap "
            f"{rolling.get('overlapSnapshots')} · {rolling.get('windowCount')} windows",
        )

    summary = result.get("summary")
    if isinstance(summary, list) and summary:
        section("Results")
        for entry in summary:
            if not isinstance(entry, dict):
                continue
            value = entry.get("value")
            detail = entry.get("detail")
            shown = f"{value}  ({detail})" if value not in (None, "") and detail else value
            add(str(entry.get("label", "")), shown)

    constraints = scenario.get("constraints")
    if isinstance(constraints, list):
        active = [c for c in constraints if isinstance(c, dict) and c.get("enabled")]
        if active:
            section("Active constraints")
            for c in active:
                add(str(c.get("label") or c.get("metric") or "constraint"), "applied")

    section("Settings")
    cp = scenario.get("carbonPrice")
    # Same fallback as the solve pipeline (network/__init__.py, results) — the
    # symbol is user-set via Settings -> currencySymbol, never assumed.
    cur = options.get("currencySymbol") or "$"
    add("Carbon price", f"{cp} {cur}/t" if cp not in (None, "") else None)
    dr = scenario.get("discountRate")
    add("Discount rate", f"{dr * 100:.1f}%" if isinstance(dr, (int, float)) else None)
    add("Currency", cur)
    if options.get("enableLoadShedding"):
        add("Load shedding", f"enabled @ {options.get('loadSheddingCost')} {cur}/MWh")

    return rows


# ── Workbook styling (display-only — never mutates cell VALUES) ────────────────
_HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
_HEADER_FONT = Font(bold=True)
# Per-cell number formats are skipped above this many cells (the wide output
# series sheets) — there they'd cost seconds and a lot of memory for little
# human benefit; those sheets still get header + freeze + width styling.
_STYLE_CELL_BUDGET = 80_000


def _number_format(value: Any) -> str | None:
    """Excel display format for a numeric cell (None ⇒ leave as-is).

    Display only — the stored value is untouched, so re-import reads the exact
    number back. ``≥1e9`` sentinels (the ``±inf`` placeholders) render compact
    scientific so a human doesn't read a literal trillion.
    """
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    av = abs(value)
    if av >= 1e9:
        return "0.0E+00"
    if float(value).is_integer():
        return "#,##0"
    if av < 1:
        return "0.0000"
    return "#,##0.00"


def _style_worksheet(ws: Any) -> None:
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 1 or max_col < 1:
        return

    # Header row: bold on a soft fill.
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    # Freeze the header row, plus the first column when it's the row key
    # (name / snapshot) so a wide sheet keeps its labels in view.
    first_header = str(ws.cell(row=1, column=1).value or "").lower()
    ws.freeze_panes = "B2" if first_header in ("name", "snapshot") else "A2"

    wide = max_col > 40  # the big output-series sheets
    sample_rows = min(max_row, 10 if wide else 40)
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        widest = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, sample_rows + 1):
            widest = max(widest, len(str(ws.cell(row=row, column=col).value or "")))
        ws.column_dimensions[letter].width = max(10, min(widest + 2, 16 if wide else 48))

    # Number formats — only on sheets small enough that touching every cell is
    # cheap. The wide series sheets keep their raw display (still round-trips).
    if max_row * max_col <= _STYLE_CELL_BUDGET:
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                fmt = _number_format(cell.value)
                if fmt is not None:
                    cell.number_format = fmt


def _style_workbook(book: Any) -> None:
    for ws in book.worksheets:
        try:
            _style_worksheet(ws)
        except Exception:  # noqa: BLE001 — styling must never fail an export
            logger.exception("Failed to style sheet %r", getattr(ws, "title", "?"))


def bundle_to_workbook(
    bundle: dict[str, Any],
    *,
    include_bundle: bool = False,
    include_meta: bool = True,
    include_model: bool = True,
    include_result: bool = True,
) -> bytes:
    """Render a run bundle to a human-readable project xlsx (bytes).

    Excel is a DERIVED export — built only on an explicit user download, never
    auto-written. The three ``include_*`` flags mirror the Export dialog's
    checkboxes and select which sheet groups land in the workbook:

    * ``include_model``  — PyPSA component sheets + input time-series + snapshots
      (the PyPSA-import-ready core).
    * ``include_result`` — solved static outputs (merged into component sheets
      when the model is included), output series sheets, result meta + plugin
      analytics.
    * ``include_meta``   — Ragnarok config sheets (``RAGNAROK_*`` in the model,
      constraints, run state, settings).

    Args:
        bundle: ``{model, scenario, options, result}`` (as stored by run_store)
            or the lighter ``{model, result}`` posted by Export Project.
        include_bundle: when True, also embed the complete bundle as chunked
            JSON (``RAGNAROK_Bundle``) so a *standalone* xlsx round-trips
            losslessly. The project *package* ships the JSON separately, so it
            leaves this off and the xlsx stays clean for Excel viewing.
        include_meta: write the Ragnarok metadata/config sheets.
        include_model: write the PyPSA input sheets.
        include_result: write the solved outputs.

    Returns:
        xlsx file bytes (readable component / series / metadata sheets).
    """
    model = bundle.get("model") or {}
    result = bundle.get("result") or {}
    scenario = bundle.get("scenario") or {}
    options = bundle.get("options") or {}
    outputs = result.get("outputs") if isinstance(result, dict) else None
    outputs = outputs or {}
    static = (outputs.get("static") or {}) if include_result else {}
    series = (outputs.get("series") or {}) if include_result else {}

    buffer = BytesIO()
    used: set[str] = set()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        wrote_any = False
        series_keys = set(series.keys())

        # 1) Model sheets (components + input temporal + config). Component
        #    sheets get their solved static outputs merged in (Result part).
        #    RAGNAROK_* config sheets inside the model belong to the Metadata
        #    part; everything else is the Model part.
        if isinstance(model, dict):
            for sheet, rows in model.items():
                if sheet in series_keys or not isinstance(rows, list):
                    continue
                is_config_sheet = str(sheet).startswith("RAGNAROK_")
                if not (include_meta if is_config_sheet else include_model):
                    continue
                if ps.component_schema(sheet) is not None:
                    rows = _merge_static_outputs(sheet, rows, static.get(sheet) or {})
                wrote_any |= _write_rows(writer, str(sheet), rows, used)

        # 2) Static outputs for components with no input rows in the workbook
        #    (also the home for ALL static outputs when the model is excluded).
        if isinstance(static, dict):
            for sheet, comp_map in static.items():
                already_written = include_model and isinstance(model, dict) and sheet in model
                if not isinstance(comp_map, dict) or already_written:
                    continue
                rows = [{"name": name, **attrs} for name, attrs in comp_map.items()]
                wrote_any |= _write_rows(writer, str(sheet), rows, used)

        # 3) Output time-series sheets (`<list>-<attr>`).
        if isinstance(series, dict):
            for key, rows in series.items():
                if isinstance(rows, list):
                    wrote_any |= _write_rows(writer, str(key), rows, used)

        # 4) Result meta + plugin analytics (Result part).
        if include_result:
            wrote_any |= _write_rows(writer, RESULT_META_SHEET, _result_meta_rows(result), used)
            plugin_rows = _plugin_analytics_rows(result.get("pluginAnalytics") if isinstance(result, dict) else None)
            _write_rows(writer, PLUGIN_ANALYTICS_SHEET, plugin_rows, used)

        # 5) Ragnarok config sheets (Metadata part).
        if include_meta:
            constraints = scenario.get("constraints") if isinstance(scenario, dict) else None
            if isinstance(constraints, list):
                _write_rows(writer, CONSTRAINTS_SHEET, constraints, used)

            run_state = _kv_rows(
                [
                    ("snapshotStart", options.get("snapshotStart")),
                    ("snapshotEnd", options.get("snapshotEnd")),
                    ("snapshotWeight", options.get("snapshotWeight")),
                    ("carbonPrice", scenario.get("carbonPrice") if isinstance(scenario, dict) else None),
                    ("forceLp", options.get("forceLp")),
                    ("activeScenarioId", options.get("scenarioLabel")),
                ]
            )
            _write_rows(writer, RUN_STATE_SHEET, run_state, used)

            settings = _kv_rows(
                [
                    ("currencySymbol", options.get("currencySymbol")),
                    ("discountRate", scenario.get("discountRate") if isinstance(scenario, dict) else None),
                    ("dateFormat", options.get("dateFormat")),
                ]
            )
            _write_rows(writer, SETTINGS_SHEET, settings, used)

        if include_bundle:
            # Embed the complete bundle as chunked JSON so even a *standalone*
            # xlsx round-trips losslessly. Off by default: the project package
            # ships the canonical JSON as its own file, keeping the xlsx clean.
            bundle_rows = [
                {"part": part, "json": chunk}
                for part, chunk in enumerate(_chunks(json.dumps(bundle, default=str)))
            ]
            wrote_any |= _write_rows(writer, BUNDLE_SHEET, bundle_rows, used)

        if not wrote_any:
            pd.DataFrame([{"info": "No data in this project."}]).to_excel(
                writer, sheet_name="info", index=False
            )

        # Human-readable landing sheet — written last, then moved to the front so
        # the workbook opens on a skimmable overview instead of the raw `network`
        # sheet. Skipped on re-import (it's in _META_SHEETS). It surfaces KPIs +
        # settings, so it belongs to the Result/Metadata parts — a pure model-only
        # export stays clean PyPSA inputs with no RAGNAROK_ sheets.
        summary_rows = _summary_rows(bundle) if (include_result or include_meta) else []
        if summary_rows and _write_rows(writer, SUMMARY_SHEET, summary_rows, used):
            book = writer.book
            sheets = book._sheets  # openpyxl keeps sheet order here
            idx = next((i for i, s in enumerate(sheets) if s.title == SUMMARY_SHEET), None)
            if idx is not None and idx != 0:
                sheets.insert(0, sheets.pop(idx))

        # Display styling (bold headers, frozen panes, widths, number formats).
        # Purely cosmetic — cell VALUES are never changed, so re-import is
        # byte-identical in meaning.
        _style_workbook(writer.book)

    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  READ  —  xlsx bytes → bundle
# ─────────────────────────────────────────────────────────────────────────────
def _df_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    """DataFrame → list-of-dicts with NaN → None (column order preserved)."""
    clean = df.where(pd.notnull(df), None)
    return clean.to_dict(orient="records")


def _kv_map(rows: list[dict[str, Any]]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for row in rows:
        key = row.get("key")
        if key is None or str(key).strip() == "":
            continue
        out[str(key)] = row.get("value")
    return out


def _reassemble_meta(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_key: dict[str, dict[int, str]] = {}
    for row in rows:
        key = row.get("key")
        if not key or not isinstance(row.get("json"), str):
            continue
        by_key.setdefault(str(key), {})[int(row.get("part") or 0)] = row["json"]
    out: dict[str, Any] = {}
    for key, parts in by_key.items():
        joined = "".join(parts[i] for i in sorted(parts))
        if not joined.strip():
            continue
        try:
            out[key] = json.loads(joined)
        except json.JSONDecodeError:
            logger.warning("Skipping malformed RAGNAROK_ResultMeta key %r", key)
    return out


def _reassemble_plugin(rows: list[dict[str, Any]]) -> dict[str, Any]:
    acc: dict[str, dict[str, Any]] = {}
    for row in rows:
        module_id = row.get("moduleId")
        if not module_id:
            continue
        entry = acc.setdefault(str(module_id), {"name": str(row.get("name") or module_id), "ui": {}, "data": {}})
        field = str(row.get("field") or "")
        if field not in ("ui", "data"):
            continue
        entry.setdefault(f"_{field}_parts", {})[int(row.get("part") or 0)] = (
            row.get("value") if isinstance(row.get("value"), str) else ""
        )
    plugin: dict[str, Any] = {}
    for module_id, entry in acc.items():
        result_entry = {"name": entry["name"], "ui": {}, "data": {}}
        for field in ("ui", "data"):
            parts = entry.get(f"_{field}_parts", {})
            joined = "".join(parts[i] for i in sorted(parts))
            if joined.strip():
                try:
                    result_entry[field] = json.loads(joined)
                except json.JSONDecodeError:
                    pass
        plugin[module_id] = result_entry
    return plugin


def workbook_to_bundle(data: bytes, filename: str = "") -> dict[str, Any]:
    """Parse a project xlsx back into a run bundle ``{model, scenario, options, result}``.

    The reverse of :func:`bundle_to_workbook`. Output static columns inside
    component sheets are split into ``result.outputs.static``; ``<list>-<attr>``
    output series sheets into ``result.outputs.series``. Derived analytics
    (summary, dispatch, …) are intentionally *not* reconstructed here — the
    frontend recomputes them from ``outputs`` when the run is opened.
    """
    excel = pd.ExcelFile(BytesIO(data), engine="openpyxl")

    # Fast path: a Ragnarok-exported workbook embeds the complete bundle as
    # chunked JSON. Read it back verbatim — the imported run is then identical
    # to the exported one, with every derived field intact (no reconstruction).
    if BUNDLE_SHEET in excel.sheet_names:
        rows = _df_rows(excel.parse(BUNDLE_SHEET))
        parts = {int(r.get("part") or 0): r.get("json") for r in rows if isinstance(r.get("json"), str)}
        joined = "".join(parts[i] for i in sorted(parts))
        if joined.strip():
            try:
                bundle = json.loads(joined)
                bundle.setdefault("options", {}).setdefault("filename", filename)
                return bundle
            except json.JSONDecodeError:
                logger.warning("RAGNAROK_Bundle JSON unreadable — falling back to sheet reconstruction")

    # Fallback: reconstruct from the readable sheets (e.g. a hand-built workbook
    # with no embedded bundle). Derived analytics are recomputed on the client.
    model: dict[str, Any] = {}
    static: dict[str, dict[str, Any]] = {}
    series: dict[str, Any] = {}
    result: dict[str, Any] = {"outputs": {"static": static, "series": series}}
    scenario: dict[str, Any] = {}
    options: dict[str, Any] = {}

    for sheet in excel.sheet_names:
        rows = _df_rows(excel.parse(sheet))

        if sheet == RESULT_META_SHEET:
            result.update(_reassemble_meta(rows))
            continue
        if sheet == CONSTRAINTS_SHEET:
            if rows:
                scenario["constraints"] = rows
            continue
        if sheet == RUN_STATE_SHEET:
            kv = _kv_map(rows)
            for key in ("snapshotStart", "snapshotEnd", "snapshotWeight", "forceLp"):
                if key in kv:
                    options[key] = kv[key]
            if "carbonPrice" in kv:
                scenario["carbonPrice"] = kv["carbonPrice"]
            if kv.get("activeScenarioId"):
                options["scenarioLabel"] = kv["activeScenarioId"]
            continue
        if sheet == SETTINGS_SHEET:
            kv = _kv_map(rows)
            if "currencySymbol" in kv:
                options["currencySymbol"] = kv["currencySymbol"]
            if "dateFormat" in kv:
                options["dateFormat"] = kv["dateFormat"]
            if "discountRate" in kv:
                scenario["discountRate"] = kv["discountRate"]
            continue
        if sheet == PLUGIN_ANALYTICS_SHEET:
            plugin = _reassemble_plugin(rows)
            if plugin:
                result["pluginAnalytics"] = plugin
            continue
        if sheet in _META_SHEETS:
            continue  # PluginConfigs / Provenance / RunHistory — not needed in the bundle

        # Output time-series sheet?
        comp_sheet, dash, attr = sheet.partition("-")
        if dash and attr in _output_series_attrs(comp_sheet):
            series[sheet] = rows
            continue

        # Static component sheet → split input vs output-static columns.
        comp = ps.component_schema(sheet)
        if comp is not None:
            out_attrs = _output_static_attrs(sheet)
            input_rows: list[dict[str, Any]] = []
            for row in rows:
                input_part: dict[str, Any] = {}
                output_part: dict[str, Any] = {}
                name = row.get("name")
                for key, value in row.items():
                    if key == "name":
                        input_part[key] = value
                    elif key in out_attrs:
                        if value is not None and value != "":
                            output_part[key] = value
                    else:
                        input_part[key] = value
                input_rows.append(input_part)
                if name is not None and output_part:
                    static.setdefault(sheet, {})[str(name)] = output_part
            model[sheet] = input_rows
        else:
            # Input temporal / config sheet — copy verbatim into the model.
            model[sheet] = rows

    options.setdefault("filename", filename)
    return {"model": model, "scenario": scenario, "options": options, "result": result}


# ─────────────────────────────────────────────────────────────────────────────
#  PACKAGE  —  a Ragnarok Project is a .zip of {<name>.json, <name>.xlsx}
# ─────────────────────────────────────────────────────────────────────────────
#
#  ``<name>.json``  — the canonical, complete bundle (lossless source of truth).
#  ``<name>.xlsx``  — the clean, human-readable workbook (inputs + outputs), for
#                     opening in Excel. NOT the re-import source.
#
#  Import reads the JSON member back verbatim, so a re-imported project is
#  byte-for-byte the exported one. (zip also compresses the JSON, so a package
#  is typically smaller than a single bundle-embedded xlsx.)

_PROJECT_SUFFIX = "_project"
# Data extensions stripped from a model filename when deriving a package name.
_LABEL_EXTENSIONS = (".xlsx", ".xls", ".nc", ".h5", ".hdf5", ".zip")


def bundle_to_package(
    bundle: dict[str, Any], base_name: str, meta: dict[str, Any] | None = None
) -> bytes:
    """Pack a bundle into a Ragnarok Project ``.zip`` — all three files.

    ``<stem>.json`` (canonical bundle), ``<stem>.meta.json`` (light sidecar, when
    provided), and ``<stem>.xlsx`` (readable workbook).
    """
    import zipfile

    stem = (base_name or "ragnarok").strip() or "ragnarok"
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{stem}.json", json.dumps(bundle, default=str))
        if meta is not None:
            zf.writestr(f"{stem}.meta.json", json.dumps(meta, default=str))
        zf.writestr(f"{stem}.xlsx", bundle_to_workbook(bundle))
    return buffer.getvalue()


def package_to_bundle(data: bytes, filename: str = "") -> dict[str, Any]:
    """Read a Ragnarok Project ``.zip`` back into a bundle (verbatim from its JSON).

    Falls back to parsing an embedded ``.xlsx`` member if the package has no
    JSON (e.g. a hand-assembled zip).
    """
    import zipfile

    with zipfile.ZipFile(BytesIO(data)) as zf:
        names = zf.namelist()
        # The canonical bundle is `<name>.json` — NOT the `<name>.meta.json`
        # sidecar (which is a different, lightweight file in the same package).
        json_name = next(
            (n for n in names if n.lower().endswith(".json") and not n.lower().endswith(".meta.json")),
            None,
        )
        if json_name is not None:
            bundle = json.loads(zf.read(json_name).decode("utf-8"))
            bundle.setdefault("options", {}).setdefault("filename", filename)
            return bundle
        xlsx_name = next((n for n in names if n.lower().endswith(".xlsx")), None)
        if xlsx_name is not None:
            return workbook_to_bundle(zf.read(xlsx_name), filename=filename)
    raise ValueError("Package contains no .json or .xlsx member.")


def _fill_derived_analytics(bundle: dict[str, Any]) -> dict[str, Any]:
    """Derive analytics server-side for a reconstructed bundle (X1).

    A bundle whose ``result`` carries outputs but no ``summary`` (the
    sheet-reconstruction path) historically forced the browser to re-derive
    everything. Fill it here via :func:`derive_results_from_outputs` — the same
    payload assembly as a fresh solve. Best-effort: on any failure (multi-period
    outputs, malformed series) the bundle is returned untouched and the client
    derivation path applies exactly as before.
    """
    result = bundle.get("result")
    if not isinstance(result, dict):
        return bundle
    outputs = result.get("outputs")
    has_summary = bool(result.get("summary"))
    if has_summary or not isinstance(outputs, dict) or not outputs.get("series"):
        return bundle
    try:
        from ..pypsa.results.derive_outputs import derive_results_from_outputs

        derived = derive_results_from_outputs(
            bundle.get("model") or {},
            outputs,
            bundle.get("scenario") or {},
            bundle.get("options") or {},
        )
        # Keep any fields the bundle already carried (meta, plugin analytics, …);
        # fill everything the derivation produced that is absent.
        for key, value in derived.items():
            result.setdefault(key, value)
        logger.info("import: analytics derived server-side for reconstructed bundle")
    except Exception as exc:  # noqa: BLE001 — never block an import on derivation
        logger.warning("import: server-side derivation skipped (%s)", exc)
    return bundle


def import_bundle_from_upload(data: bytes, filename: str) -> dict[str, Any]:
    """Parse an uploaded project file into a bundle, accepting a ``.zip`` package
    or a bare ``.xlsx`` (embedded-bundle fast path, else sheet reconstruction).

    Detection is by extension — an xlsx is *itself* a zip (it starts with the
    ``PK`` magic), so the magic bytes alone can't tell a package from a workbook.
    Reconstructed bundles (no embedded analytics) get their analytics derived
    server-side before returning (X1); failures fall back to client derivation.
    """
    name = filename.lower()
    if name.endswith(".zip"):
        return _fill_derived_analytics(package_to_bundle(data, filename))
    if name.endswith((".xlsx", ".xls")):
        return _fill_derived_analytics(workbook_to_bundle(data, filename=filename))
    # Unknown extension — try a package first, fall back to a bare workbook.
    try:
        return _fill_derived_analytics(package_to_bundle(data, filename))
    except Exception:  # noqa: BLE001
        return _fill_derived_analytics(workbook_to_bundle(data, filename=filename))


def project_basename(filename: str) -> str:
    """``<stem>_project`` from a model filename, stripping data extensions."""
    stem = filename.strip()
    for ext in _LABEL_EXTENSIONS:
        if stem.lower().endswith(ext):
            stem = stem[: -len(ext)]
            break
    stem = stem or "ragnarok"
    return stem if stem.endswith(_PROJECT_SUFFIX) else f"{stem}{_PROJECT_SUFFIX}"
