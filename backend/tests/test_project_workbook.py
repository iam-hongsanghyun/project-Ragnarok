"""Round-trip tests for the lossless project workbook (project_workbook.py).

A run bundle rendered to xlsx and parsed back must preserve the model inputs,
the solved outputs (static + series), and the result metadata — so an exported
project re-imports with nothing lost.
"""
from __future__ import annotations

from typing import Any

from backend.app import project_workbook as pw


def _sample_bundle() -> dict[str, Any]:
    return {
        "model": {
            "buses": [{"name": "n1", "v_nom": 380.0}, {"name": "n2", "v_nom": 380.0}],
            "generators": [
                # carries a stale p_nom_opt (an output) — it must be stripped from
                # the model and re-emerge only under outputs.static.
                {"name": "wind", "bus": "n1", "p_nom": 100.0, "carrier": "wind", "p_nom_opt": 999.0},
                {"name": "gas", "bus": "n2", "p_nom": 50.0, "carrier": "gas", "marginal_cost": 40.0},
            ],
            "loads": [{"name": "load1", "bus": "n1", "p_set": ""}],  # blank numeric cell
            "generators-p_max_pu": [
                {"snapshot": "2025-01-01T00:00:00", "wind": 0.8},
                {"snapshot": "2025-01-01T01:00:00", "wind": 0.6},
            ],
            "RAGNAROK_Scenarios": [{"id": "base", "label": "Base"}],
            "lines": [],  # empty — must be skipped, not crash
        },
        "scenario": {
            "discountRate": 0.07,
            "carbonPrice": 50.0,
            "constraints": [{"id": "c1", "enabled": True, "metric": "co2_cap", "value": 1000.0}],
        },
        "options": {
            "snapshotStart": 0,
            "snapshotEnd": 2,
            "snapshotWeight": 1,
            "currencySymbol": "$",
            "dateFormat": "auto",
            "filename": "demo.xlsx",
        },
        "result": {
            "outputs": {
                "static": {
                    "generators": {"wind": {"p_nom_opt": 120.0}, "gas": {"p_nom_opt": 50.0}},
                },
                "series": {
                    "generators-p": [
                        {"snapshot": "2025-01-01T00:00:00", "wind": 80.0, "gas": 10.0},
                        {"snapshot": "2025-01-01T01:00:00", "wind": 60.0, "gas": 30.0},
                    ],
                },
            },
            "runMeta": {"componentCounts": {"generators": 2, "buses": 2}},
            "pathway": {"enabled": False},
            "rolling": {"enabled": False},
            "narrative": ["Solved in 2 snapshots."],
            # Derived-only fields the frontend cards read directly — these must
            # survive the round-trip verbatim (they are NOT reconstructable from
            # the readable sheets, which is exactly why crashes appeared before).
            "summary": [{"label": "Total cost", "value": "$1,234", "detail": "system"}],
            "carrierMix": [{"carrier": "wind", "value": 140.0}],
            "dispatchSeries": [{"snapshot": "2025-01-01T00:00:00", "wind": 80.0}],
            "assetDetails": {"generators": {"wind": {"capacity": 120.0}}},
        },
    }


def test_embedded_xlsx_round_trips_verbatim() -> None:
    # With the bundle embedded, a standalone xlsx round-trips exactly: the
    # imported bundle is byte-for-byte the exported one (every derived field
    # intact) — an imported run then renders identically to a solved run.
    bundle = _sample_bundle()
    rt = pw.workbook_to_bundle(pw.bundle_to_workbook(bundle, include_bundle=True), filename="demo.xlsx")
    assert rt["result"] == bundle["result"]
    assert rt["model"] == bundle["model"]
    assert rt["scenario"] == bundle["scenario"]
    assert rt["options"] == bundle["options"]


def test_reconstruction_fallback_splits_inputs_and_outputs() -> None:
    bundle = _sample_bundle()
    # The clean (default) workbook has no embedded bundle, so parsing it back
    # exercises the readable-sheet reconstruction path.
    rt = pw.workbook_to_bundle(pw.bundle_to_workbook(bundle), filename="demo.xlsx")

    # Model inputs: non-empty sheets preserved; empty 'lines' dropped.
    assert {row["name"] for row in rt["model"]["generators"]} == {"wind", "gas"}
    # Output-static column split out of the model …
    assert all("p_nom_opt" not in row for row in rt["model"]["generators"])
    # … and re-homed under outputs.static.
    assert rt["result"]["outputs"]["static"]["generators"]["wind"]["p_nom_opt"] == 120.0
    series = rt["result"]["outputs"]["series"]["generators-p"]
    assert len(series) == 2 and series[0]["wind"] == 80.0
    assert "lines" not in rt["model"]  # empty sheet skipped
    assert rt["result"]["runMeta"]["componentCounts"] == {"generators": 2, "buses": 2}
    assert rt["scenario"]["constraints"][0]["metric"] == "co2_cap"


def test_workbook_never_uses_out_prefix_or_truncates() -> None:
    import openpyxl

    from io import BytesIO

    # The package's xlsx is the CLEAN, readable workbook — no embedded JSON sheet.
    data = pw.bundle_to_workbook(_sample_bundle())
    sheets = openpyxl.load_workbook(BytesIO(data), read_only=True).sheetnames
    assert not any(s.startswith("OUT_") for s in sheets), sheets
    assert "generators-p" in sheets  # output series, plainly named
    assert pw.BUNDLE_SHEET not in sheets  # clean by default
    assert all(len(s) <= 31 for s in sheets), sheets

    # Opt-in embedding still works for a standalone, lossless single file.
    embedded = pw.bundle_to_workbook(_sample_bundle(), include_bundle=True)
    embedded_sheets = openpyxl.load_workbook(BytesIO(embedded), read_only=True).sheetnames
    assert pw.BUNDLE_SHEET in embedded_sheets


def test_project_package_has_three_files_and_round_trips_verbatim() -> None:
    # A Ragnarok Project .zip carries ALL THREE files (bundle JSON + meta JSON +
    # xlsx); reading the bundle back is exact (every derived field intact).
    import zipfile

    from io import BytesIO

    bundle = _sample_bundle()
    meta = {"name": "north-sea-2030", "label": "North Sea", "kpis": []}
    pkg = pw.bundle_to_package(bundle, "north-sea-2030", meta=meta)

    names = sorted(zipfile.ZipFile(BytesIO(pkg)).namelist())
    assert names == ["north-sea-2030.json", "north-sea-2030.meta.json", "north-sea-2030.xlsx"]

    # The canonical bundle is read back verbatim — NOT confused with meta.json.
    rt = pw.package_to_bundle(pkg, filename="north-sea-2030.zip")
    assert rt["result"] == bundle["result"]
    assert rt["model"] == bundle["model"]
    assert rt["scenario"] == bundle["scenario"]
    assert rt["options"] == bundle["options"]


def test_import_from_upload_accepts_zip_and_xlsx() -> None:
    bundle = _sample_bundle()
    # zip package → verbatim
    zip_bundle = pw.import_bundle_from_upload(pw.bundle_to_package(bundle, "p"), "p.zip")
    assert zip_bundle["result"] == bundle["result"]
    # bare embedded xlsx → verbatim
    xlsx_bundle = pw.import_bundle_from_upload(
        pw.bundle_to_workbook(bundle, include_bundle=True), "p.xlsx"
    )
    assert xlsx_bundle["result"] == bundle["result"]


def test_project_basename() -> None:
    assert pw.project_basename("north-sea.xlsx") == "north-sea_project"
    assert pw.project_basename("north-sea.zip") == "north-sea_project"
    assert pw.project_basename("already_project.xlsx") == "already_project"
