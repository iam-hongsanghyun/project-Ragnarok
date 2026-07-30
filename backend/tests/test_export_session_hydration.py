"""Exports must carry the time-series the BROWSER does not hold.

The working model's series live in the session db and page into the grid on
demand — the React model keeps only the static/topology sheets (see
``stripSeriesSheets`` in ``App.tsx``). So every export whose payload comes from
the editor arrives series-less, and an export built from that payload alone drops
the entire temporal input: no ``loads-p_set``, no ``generators-p_max_pu``.

These tests pin the server-side hydration (``_model_with_session_series``) and
the three endpoints that depend on it, plus the ``scenario`` / ``options``
passthrough a re-import reads the run window back out of.
"""
from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path
from typing import Any

import pytest

from backend.app import main, model_store, session_store
from backend.app.models import ExportProjectPayload, RunPayload

SNAPS = ["2030-01-01T00:00:00", "2030-01-01T01:00:00"]


@pytest.fixture(autouse=True)
def _session_dir(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    monkeypatch.setattr(session_store, "SESSION_DIR", tmp_path / "session")
    return tmp_path


def _full_model() -> dict[str, list[dict[str, Any]]]:
    """What the SESSION holds: static sheets plus the time-series."""
    return {
        "buses": [{"name": "b", "v_nom": 380.0}],
        "carriers": [{"name": "gas"}],
        "snapshots": [{"snapshot": s} for s in SNAPS],
        "generators": [
            {"name": "g", "bus": "b", "carrier": "gas", "p_nom": 500.0, "marginal_cost": 10.0}
        ],
        "loads": [{"name": "l", "bus": "b", "p_set": 100.0}],
        "loads-p_set": [{"snapshot": s, "l": 100.0 + i} for i, s in enumerate(SNAPS)],
        "generators-p_max_pu": [{"snapshot": s, "g": 0.9} for s in SNAPS],
    }


def _editor_model() -> dict[str, list[dict[str, Any]]]:
    """What the BROWSER posts: the same model with every series sheet emptied."""
    model = _full_model()
    for sheet in ("loads-p_set", "generators-p_max_pu"):
        model[sheet] = []
    return model


def _load_session() -> None:
    model_store.save_model("default", _full_model(), filename="case.xlsx")


# ── the hydration itself ────────────────────────────────────────────────────

def test_hydration_fills_empty_series_sheets_from_the_session() -> None:
    _load_session()
    hydrated = main._model_with_session_series(_editor_model(), "default")
    assert hydrated["loads-p_set"] == _full_model()["loads-p_set"]
    assert hydrated["generators-p_max_pu"] == _full_model()["generators-p_max_pu"]


def test_hydration_never_overwrites_static_sheets_or_sent_series() -> None:
    _load_session()
    posted = _editor_model()
    # An unsaved static edit and a client-supplied series must both survive.
    posted["generators"][0]["p_nom"] = 777.0
    posted["loads-p_set"] = [{"snapshot": SNAPS[0], "l": 1.0}]
    hydrated = main._model_with_session_series(posted, "default")
    assert hydrated["generators"][0]["p_nom"] == 777.0
    assert hydrated["loads-p_set"] == [{"snapshot": SNAPS[0], "l": 1.0}]
    # The series the client left empty still comes from the session.
    assert hydrated["generators-p_max_pu"] == _full_model()["generators-p_max_pu"]


def test_hydration_without_a_session_is_a_passthrough() -> None:
    posted = _editor_model()
    assert main._model_with_session_series(posted, None) == posted
    # An unknown session must not raise — an export can still ship what it has.
    assert main._model_with_session_series(posted, "nope") == posted


# ── the endpoints ───────────────────────────────────────────────────────────

def _zip_bundle(body: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(io.BytesIO(body)) as zf:
        name = next(n for n in zf.namelist() if n.endswith(".json") and not n.endswith(".meta.json"))
        return json.loads(zf.read(name))


def test_export_project_zip_carries_the_session_series() -> None:
    _load_session()
    resp = main.export_project(
        ExportProjectPayload(
            model=_editor_model(),
            result={},
            scenario={"discountRate": 0.05, "carbonPrice": 25.0},
            options={"snapshotStart": 0, "snapshotEnd": 2, "snapshotWeight": 1},
            sessionId="default",
        )
    )
    bundle = _zip_bundle(resp.body)
    assert bundle["model"]["loads-p_set"] == _full_model()["loads-p_set"]
    assert bundle["model"]["generators-p_max_pu"] == _full_model()["generators-p_max_pu"]
    # scenario / options travel too — a re-import reads the run window off these,
    # and reading them off an absent `options` opened at 0 snapshots.
    assert bundle["options"]["snapshotEnd"] == 2
    assert bundle["scenario"]["carbonPrice"] == 25.0


def test_export_workbook_xlsx_carries_the_session_series() -> None:
    _load_session()
    resp = main.export_workbook(
        ExportProjectPayload(
            model=_editor_model(),
            scenario={"discountRate": 0.05},
            options={"snapshotStart": 0, "snapshotEnd": 2},
            sessionId="default",
        )
    )
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(resp.body), read_only=True)
    assert "loads-p_set" in book.sheetnames
    assert "generators-p_max_pu" in book.sheetnames
    rows = list(book["loads-p_set"].values)
    assert rows[0][:2] == ("snapshot", "l")
    assert len(rows) == 1 + len(SNAPS)


def test_netcdf_export_network_gets_the_session_series() -> None:
    _load_session()
    payload = RunPayload(
        model=_editor_model(),
        scenario={"discountRate": 0.05},
        options={"enableLoadShedding": False, "currencySymbol": "$"},
        sessionId="default",
    )
    network = main._model_payload_to_network(payload)
    # Without hydration the load is static-only and the profile is flat.
    assert list(network.loads_t.p_set["l"]) == [100.0, 101.0]
